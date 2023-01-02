import openpyxl

inv_file=openpyxl.load_workbook("inventory.xlsx")
product_list=inv_file["Sheet1"]

products_per_suplier={}
total_value_per_supplier={}
products_under_10_inv={}

for product_row in range(2,product_list.max_row + 1):
    supplier_name=product_list.cell(product_row,4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # calculating numof supplier per supplier
    if supplier_name in products_per_suplier:
       current_num_products= products_per_suplier[supplier_name]
       products_per_suplier[supplier_name]=current_num_products+1
    else:
        print("Adding a new supplier")
        products_per_suplier[supplier_name]=1

    # calculating total value
    if supplier_name in total_value_per_supplier:
        current_total=total_value_per_supplier[supplier_name]
        total_value_per_supplier[supplier_name]=current_total+inventory*price
    else:
        total_value_per_supplier[supplier_name]=inventory*price
    # logic products less then 10 inventory
    if inventory<10:
        products_under_10_inv[int(product_num)]=int(inventory)
    #add value to file
    inventory_price.value=inventory*price

inv_file.save("newfile.xlsx")

print(products_per_suplier)
print(total_value_per_supplier)
print(products_under_10_inv)